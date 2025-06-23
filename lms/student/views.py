
from django.shortcuts import render, redirect
from .models import Program, Student,State,District
from datetime import datetime
from django.contrib import messages  
from django.shortcuts import get_object_or_404
from django.http import JsonResponse

from openpyxl import load_workbook
from django.core.exceptions import ValidationError

def studenthome(request):
    return render(request,'home.html')


def add_student(request):
    states = State.objects.all()
    programs = Program.objects.all()

    if request.method == 'POST':
        regd_no = request.POST.get('regd_no')
        name = request.POST.get('name')
        email = request.POST.get('email')
        state_id = request.POST.get('state')
        district_id = request.POST.get('district')
        city = request.POST.get('city')
        prev_degree1_name = request.POST.get('prev_degree1_name')
        prev_degree1_university = request.POST.get('prev_degree1_university')
        prev_degree1_gpa = request.POST.get('prev_degree1_gpa') or None
        prev_degree2_name = request.POST.get('prev_degree2_name')
        prev_degree2_university = request.POST.get('prev_degree2_university')
        prev_degree2_gpa = request.POST.get('prev_degree2_gpa') or None
        blood_group = request.POST.get('blood_group')
        birthday = request.POST.get('birthday') or None
        batch = request.POST.get('batch')
        status = request.POST.get('status')
        program_id = request.POST.get('program')

        if Student.objects.filter(regd_no=regd_no).exists():
            messages.error(request, f"Registration number {regd_no} already exists.")
        elif Student.objects.filter(email=email).exists():
            messages.error(request, f"Email {email} is already in use.")
        else:
            program = Program.objects.get(id=program_id)
            state_name = State.objects.get(id=state_id).name
            district_name = District.objects.get(id=district_id).name

            Student.objects.create(
                regd_no=regd_no,
                name=name,
                email=email,
                state_name=state_name,
                district_name=district_name,
                city=city,
                prev_degree1_name=prev_degree1_name,
                prev_degree1_university=prev_degree1_university,
                prev_degree1_gpa=prev_degree1_gpa,
                prev_degree2_name=prev_degree2_name,
                prev_degree2_university=prev_degree2_university,
                prev_degree2_gpa=prev_degree2_gpa,
                blood_group=blood_group,
                birthday=birthday,
                batch=batch,
                status=status,
                program=program
            )
            return redirect('studenthome')

    return render(request, 'student_form.html', {
        'states': states,
        #'districts':district,
        'programs': programs,
        'gpa_choices': [round(i * 0.1, 2) for i in range(0, 101)],
        'current_year': datetime.now().year,
        'blood_groups': Student.BLOOD_GROUP_CHOICES,
        'status': Student.STATUS_CHOICES,
    })




def update_student(request, regd_no):
    student = get_object_or_404(Student, regd_no=regd_no)
    programs = Program.objects.all()

    if request.method == 'POST':
        student.name = request.POST.get('name')
        student.email = request.POST.get('email')
        state_id = request.POST.get('state')
        district_id = request.POST.get('district')
        student.state_name = State.objects.get(id=state_id).name
        student.district_name = District.objects.get(id=district_id).name
        student.city = request.POST.get('city')
        student.prev_degree1_name = request.POST.get('prev_degree1_name')
        student.prev_degree1_university = request.POST.get('prev_degree1_university')
        student.prev_degree1_gpa = request.POST.get('prev_degree1_gpa') or None
        student.prev_degree2_name = request.POST.get('prev_degree2_name')
        student.prev_degree2_university = request.POST.get('prev_degree2_university')
        student.prev_degree2_gpa = request.POST.get('prev_degree2_gpa') or None
        student.blood_group = request.POST.get('blood_group')
        student.birthday = request.POST.get('birthday') or None
        student.batch = request.POST.get('batch')
        student.status = request.POST.get('status')

        program_id = request.POST.get('program')
        student.program = Program.objects.get(id=program_id)

        student.save()
        messages.success(request, "Student details updated successfully.")
        return redirect('studenthome')

    return render(request, 'student_update_form.html', {
        'student': student,
        'programs': programs,
        'blood_groups': Student.BLOOD_GROUP_CHOICES,
        'status_choices': Student.STATUS_CHOICES
    })


def get_states(request):
    states = State.objects.all().values('id', 'name')
    return JsonResponse(list(states), safe=False)

def get_districts_by_state(request, state_id):
    districts = District.objects.filter(state_id=state_id).values('id', 'name')
    return JsonResponse(list(districts), safe=False)









def bulk_upload_students(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        wb = load_workbook(filename=excel_file)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        expected_headers = [
            "regd_no", "name", "email", "state_name", "district_name","city", "prev_degree1_name",
            "prev_degree1_university", "prev_degree1_gpa", "prev_degree2_name",
            "prev_degree2_university", "prev_degree2_gpa", "blood_group", "birthday",
            "program", "batch", "status"
        ]

        if headers != expected_headers:
            messages.error(request, "Excel format is incorrect. Expected headers: " + ", ".join(expected_headers))
            return redirect('studenthome')

        success_count = 0
        error_entries = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))

            try:
                # Validate required fields
                if not data['regd_no'] or not data['email'] or not data['name'] or not data['program']:
                    raise ValueError("Missing required fields.")

                # Check for duplicates
                if Student.objects.filter(regd_no=data['regd_no']).exists():
                    raise ValueError(f"Registration number {data['regd_no']} already exists.")
                if Student.objects.filter(email=data['email']).exists():
                    raise ValueError(f"Email {data['email']} already exists.")

                # Validate blood group
                if data['blood_group'] and data['blood_group'] not in dict(Student.BLOOD_GROUP_CHOICES):
                    raise ValueError(f"Invalid blood group: {data['blood_group']}")

                # Validate status
                if data['status'] and data['status'] not in dict(Student.STATUS_CHOICES):
                    raise ValueError(f"Invalid status: {data['status']}")

                # Validate and fetch Program
                try:
                    program = Program.objects.get(name__iexact=data['program'])
                except Program.DoesNotExist:
                    raise ValueError(f"Program '{data['program']}' does not exist.")

                # Parse birthday (expected DD-MM-YYYY)
                birthday = None
                if data['birthday']:
                    try:
                        birthday = datetime.strptime(data['birthday'], '%d-%m-%Y').date()
                    except ValueError:
                        raise ValueError("Birthday format should be DD-MM-YYYY")

                # Create student
                Student.objects.create(
                    regd_no=data['regd_no'],
                    name=data['name'],
                    email=data['email'],
                    state_name=data['state_name'] or '',
                    district_name=data.get('district_name') or '',
                    city=data.get('city') or '',
                    prev_degree1_name=data.get('prev_degree1_name') or '',
                    prev_degree1_university=data.get('prev_degree1_university') or '',
                    prev_degree1_gpa=data.get('prev_degree1_gpa') or None,
                    prev_degree2_name=data.get('prev_degree2_name') or '',
                    prev_degree2_university=data.get('prev_degree2_university') or '',
                    prev_degree2_gpa=data.get('prev_degree2_gpa') or None,
                    blood_group=data.get('blood_group'),
                    birthday=birthday,
                    program=program,
                    batch=data.get('batch') or datetime.now().year,
                    status=data.get('status') or 'active',
                    created_at=datetime.now(),
                    updated_at=datetime.now()
                )
                success_count += 1

            except Exception as e:
                error_entries.append(f"{data.get('regd_no')} - {e}")

        # Final message
        if error_entries:
            messages.warning(request, f"{success_count} students uploaded successfully. {len(error_entries)} failed.")
            messages.error(request, "\n".join(error_entries))
        else:
            messages.success(request, f"All {success_count} students uploaded successfully.")

        return redirect('studenthome')


