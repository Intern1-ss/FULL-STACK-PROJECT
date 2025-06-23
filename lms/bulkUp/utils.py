from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from django.http import HttpResponse
from DB.models import Student,Program

def generate_excel_template(columns):
    df = pd.DataFrame(columns=columns)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=template.xlsx'
    df.to_excel(response, index=False)
    return response


import pandas as pd
def read_excel_and_return_dataframe(file):
    try:
        return pd.read_excel(file)
    except Exception as e:
        return None
    


def generate_excel_template_student(columns):
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Template"

    # Add header row
    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Create dropdown validations

    # 1. Blood Group Dropdown
    blood_groups = [choice[0] for choice in Student.BLOOD_GROUP_CHOICES]
    bg_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(blood_groups)}"',
        allow_blank=True
    )
    ws.add_data_validation(bg_validation)
    bg_col = columns.index("blood_group") + 1
    bg_validation.add(f"{ws.cell(row=2, column=bg_col).column_letter}2:{ws.cell(row=1000, column=bg_col).column_letter}1000")

    # 2. Status Dropdown
    status_values = [choice[0] for choice in Student.STATUS_CHOICES]
    status_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(status_values)}"',
        allow_blank=True
    )
    ws.add_data_validation(status_validation)
    status_col = columns.index("status") + 1
    status_validation.add(f"{ws.cell(row=2, column=status_col).column_letter}2:{ws.cell(row=1000, column=status_col).column_letter}1000")

    # 3. Program Code Dropdown (from DB)
    programs = Program.objects.values_list('code', flat=True)
    program_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(programs)}"',
        allow_blank=True
    )
    ws.add_data_validation(program_validation)
    program_col = columns.index("program_code") + 1
    program_validation.add(f"{ws.cell(row=2, column=program_col).column_letter}2:{ws.cell(row=1000, column=program_col).column_letter}1000")

    # 4. Gender
    gender_values = ["Male", "Female", "Other"]
    gender_validation = DataValidation(type="list", formula1=f'"{",".join(gender_values)}"', allow_blank=True)
    col = columns.index("gender") + 1
    ws.add_data_validation(gender_validation)
    gender_validation.add(f"{ws.cell(row=2, column=col).column_letter}2:{ws.cell(row=1000, column=col).column_letter}1000")
    
    # 5. Optional: Set birthday column format
    if 'birthday' in columns:
        birthday_col = columns.index('birthday') + 1
        for row in range(2, 1001):
            ws.cell(row=row, column=birthday_col).number_format = 'DD-MM-YYYY'

    # Save to HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=student_template.xlsx'
    wb.save(response)
    return response

#for the fields like gpa 
from decimal import Decimal, InvalidOperation
def safe_decimal(value):
    """Convert value to Decimal or return None if invalid or NaN."""
    try:
        if pd.isna(value) or value == "":
            return None
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None   




    

 #for faculty
from DB.models import Campus,Department

def generate_excel_template_faculty(columns):
    wb = Workbook()
    ws = wb.active
    ws.append(columns)

    # Dropdown for campus_id
    if 'campus_id' in columns:
        campus_list = Campus.objects.all()
        campus_values = [campus.id for campus in campus_list]
        if campus_values:
            campus_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(campus_values)}"',
                allow_blank=True
            )
            ws.add_data_validation(campus_validation)
            campus_col = columns.index("campus_id") + 1
            col_letter = ws.cell(row=1, column=campus_col).column_letter
            campus_validation.add(f"{col_letter}2:{col_letter}1000")

    # Dropdown for department_name
    if 'department_name' in columns:
        dept_list = Department.objects.all()
        dept_values = [dept.dept_name.replace(",", ";") for dept in dept_list if dept.dept_name]
        if dept_values:
            dept_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(dept_values)}"',
                allow_blank=True
            )
            ws.add_data_validation(dept_validation)
            dept_col = columns.index("department_name") + 1
            col_letter = ws.cell(row=1, column=dept_col).column_letter
            dept_validation.add(f"{col_letter}2:{col_letter}1000")

    # Format DOB column
    if 'dob' in columns:
        dob_col = columns.index('dob') + 1
        for row in range(2, 1001):
            ws.cell(row=row, column=dob_col).number_format = 'DD-MM-YYYY'
    #  gender Dropdown
    gender_values = ["Male", "Female", "Other"]
    gender_validation = DataValidation(type="list", formula1=f'"{",".join(gender_values)}"', allow_blank=True)
    col = columns.index("gender") + 1
    ws.add_data_validation(gender_validation)
    gender_validation.add(f"{ws.cell(row=2, column=col).column_letter}2:{ws.cell(row=1000, column=col).column_letter}1000")
    
    #  Status Dropdown
    status_values = ['active','inactive','on leave']
    status_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(status_values)}"',
        allow_blank=True
    )
    ws.add_data_validation(status_validation)
    status_col = columns.index("status") + 1
    status_validation.add(f"{ws.cell(row=2, column=status_col).column_letter}2:{ws.cell(row=1000, column=status_col).column_letter}1000")


    # Prepare Excel download response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=faculty_template.xlsx'
    wb.save(response)
    return response


# for department
from DB.models import Faculty
def generate_excel_template_department(columns):
    wb = Workbook()
    ws = wb.active
    ws.append(columns)

    # Dropdown for HOD (faculty names)
    if 'hod_name' in columns:
        faculty_list = Faculty.objects.all()
        faculty_names = [faculty.name.replace(",", ";") for faculty in faculty_list if faculty.name]
        if faculty_names:
            # Excel formula allows max 255 chars, use only if names are short
            hod_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(faculty_names)}"',
                allow_blank=True
            )
            ws.add_data_validation(hod_validation)
            hod_col = columns.index("hod_name") + 1
            hod_validation.add(
                f"{ws.cell(row=2, column=hod_col).column_letter}2:{ws.cell(row=1000, column=hod_col).column_letter}1000"
            )

    # Prepare Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=department_template.xlsx'
    wb.save(response)
    return response



# for program
# from DB.models import Program
def generate_excel_template_program(columns):
    # columns = ['name', 'code', 'department_name', 'duration_years']
    wb = Workbook()
    ws = wb.active
    ws.append(columns)

    # Department name dropdown
    
    dept_names = [dept.dept_name.replace(',', ';') for dept in Department.objects.all()]
    
    if dept_names:
        formula = f'"{",".join(dept_names)}"'
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        ws.add_data_validation(validation)
        col_letter = ws.cell(row=1, column=columns.index("department_name") + 1).column_letter
        validation.add(f"{col_letter}2:{col_letter}1000")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=program_template.xlsx'
    wb.save(response)
    return response